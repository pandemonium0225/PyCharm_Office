import os
from openpyxl import load_workbook
from openpyxl import drawing
from win32com import client
import win32api
import re
import shutil


def renaming_schedule():
    schedule_location = input("please paste the directory you want to rename...")
    schedule_folder = "schedules"
    folder = os.path.join(schedule_location, schedule_folder)
    os.makedirs(folder, exist_ok=True)
    for path,_,files in os.walk(schedule_location):
        for file in files:
            filepath = os.path.join(path,file)
            wb = load_workbook(filepath,data_only=True)
            sheet = wb.worksheets[3]
            img = drawing.image.Image(r"C:\Users\sebein\Desktop\makeup schedule\Sean_Stamp\sean.png")
            sheet.add_image(img,"d61")
            client = sheet.cell(8, 4).value
            schedule_name = sheet.cell(10,4).value
            brand = sheet['D9'].value
            media = sheet['C18'].value
            split_media = media.split("-")[0].lower()
            phase2 = schedule_name + "_P2_" + brand + ".xlsx"
            phase1 = schedule_name + "_P1_" + brand + ".xlsx"
            gam = schedule_name + "_GoogleAsia" + ".xlsx"
            # wb.save(filepath)
            if ("TW9098") in client:
                print("renaming {} to {}".format(file,phase2))
                wb.save(os.path.join(folder, phase2))
                # os.rename(filepath,os.path.join(path, phase2))
            elif ("TW8081") in client:
                print("renaming {} to {}".format(file,phase2))
                wb.save(os.path.join(folder, phase2))
                # print("renaming {} to {}".format(file,phase1))
                # wb.save(os.path.join(folder, phase1))
                # os.rename(filepath,os.path.join(path, phase1))
            elif "GOOASI" in client:
                print("renaming {} to {}".format(file,gam))
                wb.save(os.path.join(folder,gam))
            elif "TW1713" in client:
                print("renaming {} to {}".format(file, phase1))
                wb.save(os.path.join(folder,phase1))
            elif "HK9067" in client:
                print("renaming {} to {}".format(file, phase1))
                wb.save(os.path.join(folder, phase1))
            elif "TW3367" in client:
                print("renaming {} to {}".format(file, phase2))
                wb.save(os.path.join(folder, phase2))
            else:
                print("Non_Apex Campaign, passed")
            wb.close()


def exceltopdf(doc):
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = 0

    wb = excel.Workbooks.Open(doc)
    ws = wb.Worksheets[1]

    try:
        wb.SaveAs(doc[:-5], FileFormat=57)
    except Exception as e:
        print("Failed to convert")
        print(e)
    finally:
        wb.Close()
        excel.Quit()


def convert_to_pdf():
    schedule_location = input("Please input the location where the schedule excel file exists")
    for path, _, files in os.walk(schedule_location):
        for file in files:
            filepath = os.path.join(path, file)
            try:
                exceltopdf(filepath)
            except Exception as e:
                print(e)


def get_jobno_pair():
    job_pair = {}
    folder_name_list = []
    schedule_location = input("Please input the location where the schedule excel file exists")
    for path, dirs, files in os.walk(schedule_location):
        for file in files:
            try:
                name = re.search(r"(APX.*?)\.\w{3,4}",file).group(1)
                job = name.split('_')[0]
                client = name.split('_')[2]
                media = re.search(r'APX(\w{2})\d{4}',name)
                folder_name = client + "_" + media.group(1)
                folder_name_list.append(folder_name)
                if client != "General":
                    job_pair[job] = client
            except Exception as e:
                print(file,e)
    return job_pair, folder_name_list, schedule_location


def create_folders():
    job_pair, folder_name_list, schedule_location = get_jobno_pair()
    folder_name_list = [x for x in folder_name_list if any(["General" not in x, x.endswith(".pdf")])]
    folder_name_list = list(dict.fromkeys(folder_name_list))
    print(folder_name_list)
    print(job_pair)
    folder_location = input("please input the folder location!!!")
    for n in range(len(folder_name_list)):
        os.mkdir(folder_location + '\\' + folder_name_list[n])
    for root,dirs,files in os.walk(schedule_location):
        for file in files:
            name = re.search(r"(APX.*?)\.\w{3,4}", file).group(1)
            job = name.split('_')[0]
            # client = name.split('_')[2]
            media = re.search(r'APX(\w{2})\d{4}', name).group(1)
            client = job_pair[job]
            if file.endswith('.pdf'):
                print(job_pair[job])
                print(root + "\\" + file)
                print(root + "\\" + client + "_" + media + "\\" + file)
                shutil.copy(root + "\\" + file, folder_location + "\\" + client + "_" + media + "\\" + file)


if __name__ == '__main__':
    renaming_schedule()
    convert_to_pdf()
    create_folders()