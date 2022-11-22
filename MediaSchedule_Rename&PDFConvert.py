import os
from openpyxl import load_workbook
from openpyxl import drawing
from functools import wraps
import time

def fn_timer(function):
    '''
    函数计时装饰器
    :param function: 函数对象
    :return: 装饰器
    '''
    @wraps(function)
    def function_timer(*args,**kwargs):
        # 起始时间
        t0 = time.time()
        # 调用函数
        result = function(*args,**kwargs)
        # 结束时间
        t1 = time.time()
        # 打印函数耗时
        print('[finished function:{func_name} in {time:.2f}s]'.format(func_name = function.__name__,time = t1 - t0))
        return result
    return function_timer

@fn_timer
def schedule_renaming():
    location = input("Please input or paste the directory under which the file you want to rename") + "\\"
    for path, dirs, files in os.walk(location):
        for file in files:
            filepath = os.path.join(path,file)
            wb = load_workbook(filepath,data_only=True)
            sheet = wb.worksheets[3]
            # img = drawing.image.Image(r"D:\OfficeFunction\office_function\Images\sean.png")
            # sheet.add_image(img,"d61")
            schedule_name = sheet.cell(10,4).value
            client = sheet.cell(8, 4).value
            phase2 = client + schedule_name + "_P2" + ".xlsx"
            phase1 = client + schedule_name + "_p1" + ".xlsx"
            wb.save(filepath)
            if ("TW9098") in client:
                print("renaming {} to {}".format(file, phase2))
                os.rename(filepath, os.path.join(path, phase2))
            elif ("TW8081") in client:
                print("renaming {} to {}".format(file, phase2))
                os.rename(filepath, os.path.join(path, phase2))
            elif "TW1713" in client:
                print("renaming {} to {}".format(file, phase1))
                os.rename(filepath, os.path.join(path, phase1))
            else:
                print("Non_Apex Campaign, passed")
            wb.close()


@fn_timer
def schedule_to_PDF():
    from win32com import client
    import win32api

    location = input("Please input or paste the directory under which the file you want to rename") + "\\"

    def exceltopdf(doc):
        excel = client.DispatchEx("Excel.Application")
        excel.Visible = 0

        wb = excel.Workbooks.Open(doc)
        ws = wb.Worksheets[1]

        try:
            wb.SaveAs(doc[:-5], FileFormat=57)
        except Exception as e:
            print("Failed to convert")
            print(e)
        finally:
            wb.Close()
            excel.Quit()

    for path, _, files in os.walk(location):
        for file in files:
            filepath = os.path.join(path, file)
            try:
                exceltopdf(filepath)
            except Exception as e:
                print("Conversion not completed because of ....", e)


if __name__ == "__main__":
    schedule_renaming()
    schedule_to_PDF()